import numpy as np
import cv2
import random
import openpyxl



''' ↓ 저장할 엑셀 파일 이름, 원점 그리드의 GPS 좌표, 발생시킬 사건 개수 입력'''
FILE_NAME = "crime_data_x"
O_LATITUDE = 37.5727023
O_LONGTITUDE = 126.961014
CASE_NUM = 4300

LAT_GAP = 0.0008914
LONG_GAP = 0.0011314
WIDTH = 59
HEIGHT = 32

CRIME_KEY = {
    '절도': 0.5089,
    '폭행': 0.9530,
    '살인': 0.9535,
    '강간': 0.9986,
    '강도': 1.0000
}


def gps2grid(lat, long):
    x = (O_LATITUDE - lat) / LAT_GAP
    y = (long - O_LONGTITUDE) / LONG_GAP
    return [x, y]

# 그리드별 좌표, 사건 별 그리드 번호
# 그리드 이미지를 배열로 변환
img = cv2.imread("grid.png", 0)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)


# 배경이 아닌 그리드의 좌표만 배열에 따로 저장 [위도, 경도]
grids = []
result = np.zeros((HEIGHT, WIDTH, 3), dtype=np.uint8)
for i in range(WIDTH):
    for j in range(HEIGHT):
        if img[j][i] == 255:
            result[j][i] = 0
        else:
            result[j][i] = 150
            grids.append([j, i])


# 엑셀 파일 생성
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = "Type"
sheet['B1'] = "Date-Time"
sheet['C1'] = "Latitude"
sheet['D1'] = "Longitude"
sheet['E1'] = "Grid"


# 정한 개수 만큼 범죄사례를 생성
for i in range(2, CASE_NUM + 2):
    # 범죄유형
    n = random.random()
    if n < CRIME_KEY['절도']:
        type = '절도'
    elif n < CRIME_KEY['폭행']:
        type = '폭행'
    elif n < CRIME_KEY['살인']:
        type = '살인'
    elif n < CRIME_KEY['강간']:
        type = '강간'
    elif n < CRIME_KEY['강도']:
        type = '강도'
    sheet.cell(row=i, column=1).value = type

    # 범죄시간
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row=i, column=2).value = date + " " + time

    # 범죄지점 : 그리드 좌표 중 랜덤한 구역을 정하고, 그 구역 안에서 랜덤생성
    idx = int(random.uniform(0, len(grids)))
    x, y = grids[idx]
    gx, gy = O_LATITUDE - x * LAT_GAP, O_LONGTITUDE + y * LONG_GAP
    sheet.cell(row=i, column=3).value = random.uniform(gx - LAT_GAP, gx)
    sheet.cell(row=i, column=4).value = random.uniform(gy, gy + LONG_GAP)
    sheet.cell(row=i, column=5).value = idx
    result[x][y] += 20

# 파일 저장
wb.save(FILE_NAME + '.xlsx')


# 배열이 이미지와 같게 구성되었는지 확인
result = cv2.resize(dsize=(WIDTH*20, HEIGHT*20), src=result, interpolation=cv2.INTER_NEAREST)
cv2.imshow("res", result)
cv2.waitKey()


