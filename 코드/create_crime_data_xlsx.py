############ create crime data ############
import openpyxl
import random

# create excel
wb = openpyxl.Workbook()

# open sheet
sheet = wb.active

# save label
sheet['A1'] = "Type"
sheet['B1'] = "Date-Time"
sheet['C1'] = "Latitude"
sheet['D1'] = "Longitude"

############################# random #############################
# create data (1050)
for i in range(2, 1052):
    sheet.cell(row = i, column = 1).value = "Theft"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.5546645, 37.56886886)
    sheet.cell(row = i, column = 4).value = random.uniform(126.96777894, 127.0176931)

# create data (950)
for i in range(1052, 2002):
    sheet.cell(row = i, column = 1).value = "Assault"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.5546645, 37.56886886)
    sheet.cell(row = i, column = 4).value = random.uniform(126.96777894, 127.0176931)

# create data (100)
for i in range(2002, 2102):
    sheet.cell(row = i, column = 1).value = "Sexaul Harassment"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.5546645, 37.56886886)
    sheet.cell(row = i, column = 4).value = random.uniform(126.96777894, 127.0176931)

## create data (6)
#for i in range(4292, 4298):
#    sheet.cell(row = i, column = 1).value = "Robber"
#    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
#    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
#    sheet.cell(row = i, column = 2).value = date + " " + time
#    sheet.cell(row = i, column = 3).value = random.uniform(37.5537680, 37.5688683)
#    sheet.cell(row = i, column = 4).value = random.uniform(126.9677753, 127.0176953)
#
## create data (2)
#for i in range(4298, 4300):
#    sheet.cell(row = i, column = 1).value = "Murder"
#    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
#    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
#    sheet.cell(row = i, column = 2).value = date + " " + time
#    sheet.cell(row = i, column = 3).value = random.uniform(37.5537680, 37.5688683)
#    sheet.cell(row = i, column = 4).value = random.uniform(126.9677753, 127.0176953)

# save excel(workbook)
wb.save('crime_data_red.xlsx')