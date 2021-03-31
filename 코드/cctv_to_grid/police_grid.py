import numpy as np
import cv2
import openpyxl

''' ↓ 저장할 엑셀 파일 이름 입력'''
FILE_NAME = "Police_grid"

O_LATITUDE = 37.57244364
O_LONGTITUDE = 126.96095890
LAT_GAP = 0.000896694
LONG_GAP = 0.001138112

WIDTH = 59
HEIGHT = 32

# 그리드 이미지를 배열로 변환
img = cv2.imread("grid.png", 0)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)

# 배경이 아닌 그리드의 좌표만 배열에 따로 저장 [위도, 경도]
grids = []
result = np.zeros((HEIGHT, WIDTH, 3), dtype=np.uint8)
for i in range(WIDTH):
    for j in range(HEIGHT):
        if img[j][i] != 255:
            grids.append([j, i])

# 기존 cctv 엑셀 파일 불러오기
latitude_num, longitude_num = 4, 5
load_wb = openpyxl.load_workbook("Police.xlsx", data_only=True)
load_ws = load_wb['Police']

# 새로운 엑셀 파일 생성(연도별 시트 생성)
wb = openpyxl.Workbook()
sheet = wb.create_sheet('Police_grid')
sheet['A1'] = "a"
sheet['B1'] = "b"
sheet['C1'] = "c"
sheet['D1'] = "Latitude"
sheet['E1'] = "Longitude"
sheet['F1'] = "f"

# 기존 내용 옮겨 적기 (grid제외 정보)
for i in range(1, len(load_ws['A'])):
    for j in range(1, 7):
        sheet.cell(row=i, column=j).value = load_ws.cell(i, j).value

# 위도 경도에 따른 grid 번호 부여
sheet.cell(row=1, column=7).value = "Grid"
for i in range(0, len(grids)):
    x, y = grids[i]
    gx, gy = O_LATITUDE - x * LAT_GAP, O_LONGTITUDE + y * LONG_GAP # 해당 그리드에서 좌상단 좌표
    for j in range(2, len(load_ws['A'])):
        load_lat, load_long = load_ws.cell(j, latitude_num).value, load_ws.cell(j, longitude_num).value
        if load_lat <= gx and load_lat >= gx - LAT_GAP and load_long >= gy and load_long <= gy + LONG_GAP:
            sheet.cell(row=j, column=7).value = i

# 파일 저장
wb.remove(wb['Sheet'])
wb.save(FILE_NAME + '.xlsx')
print('Excel file created!')


