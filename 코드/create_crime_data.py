############ create crime data ############
import openpyxl
import random

# create excel
wb = openpyxl.Workbook()

# open sheet
sheet = wb.active

# save label
sheet['A1'] = "Type"
sheet['B1'] = "Date-Time"
sheet['C1'] = "Latitude"
sheet['D1'] = "Longitude"

# create data (2200)
for i in range(2, 2202):
    sheet.cell(row = i, column = 1).value = "절도"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.554842, 37.567894)
    sheet.cell(row = i, column = 4).value = random.uniform(126.968953, 127.016778)

# create data (1900)
for i in range(2202, 4102):
    sheet.cell(row = i, column = 1).value = "폭행"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.554842, 37.567894)
    sheet.cell(row = i, column = 4).value = random.uniform(126.968953, 127.016778)

# create data (190)
for i in range(4102, 4292):
    sheet.cell(row = i, column = 1).value = "강간"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.554842, 37.567894)
    sheet.cell(row = i, column = 4).value = random.uniform(126.968953, 127.016778)

# create data (6)
for i in range(4292, 4298):
    sheet.cell(row = i, column = 1).value = "강도"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.554842, 37.567894)
    sheet.cell(row = i, column = 4).value = random.uniform(126.968953, 127.016778)

# create data (2)
for i in range(4298, 4300):
    sheet.cell(row = i, column = 1).value = "살인"
    date = "2019-" + str(random.randrange(1, 13)) + "-" + str(random.randrange(1, 31))
    time = str(random.randrange(0, 24)) + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
    sheet.cell(row = i, column = 2).value = date + " " + time
    sheet.cell(row = i, column = 3).value = random.uniform(37.554842, 37.567894)
    sheet.cell(row = i, column = 4).value = random.uniform(126.968953, 127.016778)

# save excel(workbook)
wb.save('crime_data.xlsx')