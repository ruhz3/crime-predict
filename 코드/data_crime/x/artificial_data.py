'''
- artificial_data.py
 : 지도 화면을 클릭해 색칠된 그리드에만 데이터를 생성

* 사용방법
1. 좌표를 화면에 찍고
2. '1' + 'q' 키를 누르면
3. 색칠된 그리드 내에서만 범죄 데이터 생성
'''

import numpy as np
import cv2
import random
import openpyxl
from const_data import *


''' ↓ 저장할 엑셀 파일 이름, 연도별 경향성 데이터 수 입력'''
FILE_NAME = "crime_data_a"
CRIME_TOTAL = [1000, 1000, 1000, 1000, 1000, 1000]  # 2014년 - 2019년


WIDTH = 59
HEIGHT = 32

CRIME_KEY = {0: 'Theft', 1: 'Violence', 2: 'Murder', 3: 'Rape', 4: 'Robbery'}
YEAR_KEY = {0: '2014', 1: '2015', 2: '2016', 3: '2017', 4: '2018', 5: '2019'}

# 그리드 이미지를 배열로 변환
img = cv2.imread("grid.png", 0)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)
grids = []

# <editor-fold desc="나타나는 화면에 찍은 좌표에만 사례 랜덤하게 생성">
# 마우스 클릭 시 그리드 좌표 출력하는 Callback 등록
def select_grid(event, x, y, flags, param):
    if event == cv2.EVENT_LBUTTONDBLCLK:
        x = int(x / 10)
        y = int(y / 10)
        img[y][x] = 0
        grids.append([y, x])
        print(f'{[y, x]}')
cv2.namedWindow("viewer")
cv2.setMouseCallback("viewer", select_grid)

# 렌더링
while cv2.waitKey(1) != ord('q'):
    view = np.zeros((HEIGHT, WIDTH), dtype=np.uint8)
    for x in range(WIDTH):
        for y in range(HEIGHT):
            view[y][x] = img[y][x]
    view = cv2.resize(view, (WIDTH*10, HEIGHT*10), interpolation=cv2.INTER_NEAREST)
    cv2.imshow("viewer", view)
# </editor-fold>
print("Grids selected!")

# 엑셀 파일 생성(연도별 시트 생성)
wb = openpyxl.Workbook()
for year in range(6):
    print(f'{YEAR_KEY[year]}...')
    sheet = wb.create_sheet(f'{YEAR_KEY[year]}')
    sheet['A1'] = "Type"
    sheet['B1'] = "Date-Time"
    sheet['C1'] = "Latitude"
    sheet['D1'] = "Longitude"
    sheet['E1'] = "Grid"
    case_num = CRIME_TOTAL[year]

    # 통계자료에 맞춰 범죄사례를 생성
    for i in range(2, case_num + 2):
        # 범죄유형
        n = random.random()
        for c in range(5):
            if n < CRIME_STATS[year][c]:
                sheet.cell(row=i, column=1).value = f'{CRIME_KEY[c]}'
                crime = c
                break

        # 범죄날짜 및 시간
        month = random.randrange(1, 13)
        day = random.randrange(1, MONTH_DAYS[month] + 1)
        if month < 10:
            str_month = f'0{month}'
        else:
            str_month = f'{month}'
        if day < 10:
            str_day = f'0{day}'
        else:
            str_day = f'{day}'
        date = f'{YEAR_KEY[year]}-{str_month}-{str_day}'

        n = random.random()
        for k in range(8):
            if n < CRIME_TIME[crime][k]:
                hour = random.randrange(3*k, 3*(k+1))
                break
        if hour < 10:
            str_hour = f'0{hour}'
        else:
            str_hour = f'{hour}'
        time = str_hour + ":00:00"
        sheet.cell(row=i, column=2).value = date + " " + time

        # 범죄지점 : 그리드를 랜덤으로 선택하고, 그리드 내부에서 랜덤생성
        idx = int(random.uniform(0, len(grids)))
        x, y = grids[idx]
        gx, gy = O_LATITUDE - x * LAT_GAP, O_LONGTITUDE + y * LONG_GAP
        sheet.cell(row=i, column=3).value = random.uniform(gx - LAT_GAP, gx)
        sheet.cell(row=i, column=4).value = random.uniform(gy, gy + LONG_GAP)
        sheet.cell(row=i, column=5).value = idx


# 파일 저장
wb.remove(wb['Sheet'])
wb.save(FILE_NAME + '.xlsx')
print('Excel file created!')



