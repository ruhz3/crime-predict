import numpy as np
import cv2
import openpyxl
import random


WIDTH = 59
HEIGHT = 32

DONG_DICT = {
    0: '중림동', 1: '소공동', 2: '회현동', 3: '명동',  4: '필동', 5: '을지로동',
    6: '장충동', 7: '광희동', 8: '다산동', 9: '약수동', 10: '신당동', 11: '청구동',
    12: '동화동', 13: '신당5동', 14: '황학동'
}

COLOR_DICT = {
    0: [54, 182, 229], 1: [176, 228, 239], 2: [21, 0, 136], 3: [127, 127, 127], 4: [14, 201, 255], 5: [29, 230, 181],
    6: [164, 73, 163], 7: [87, 122, 185], 8: [232, 162, 0], 9: [201, 174, 255], 10: [76, 177, 34], 11: [0, 242, 255],
    12: [39, 127, 255], 13: [204, 72, 63], 14: [36, 28, 237]
}

TIME_DICT = {
    0: 'Morning',
    1: 'Afternoon',
    2: 'Evening',
    3: 'Midnight'
}

FLOAT_ARR = [
    [12170,	32230,	41700,	19220],
    [32870,	87020,	112600,	51900],
    [18260,	48340,	62550,	28830],
    [24350,	64450,	83410,	38450],
    [29220,	77350,	100090,	46140],
    [30440,	80570,	104260,	48060],
    [10350,	27390,	35450,	16340],
    [48700,	128920,	166820,	76900],
    [91320,	241720,	312790,	144190],
    [111400, 294900, 381610, 175910],
    [90710,	240110,	310710,	143220],
    [59050,	156310,	202270,	93240],
    [68790,	182100,	235640,	108620],
    [73660,	194990,	252320,	116310],
    [51740,	136970,	177250,	81700]
]

# 그리드 이미지를 배열로 변환
img = cv2.imread("dong.png", 1)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)
grids = []
for x in range(WIDTH):
    for y in range(HEIGHT):
        if img[y][x][0] != 255 or img[y][x][1] != 255 or img[y][x][2] != 255:
            grids.append([y, x])

# 그리드를 동별로 분리
grids_dong = []
for i in range(15):
    new = []
    for x in range(WIDTH):
        for y in range(HEIGHT):
            if img[y][x][0] == COLOR_DICT[i][0] and img[y][x][1] == COLOR_DICT[i][1] and img[y][x][2] == COLOR_DICT[i][2]:
                new.append([y, x])
    grids_dong.append(new)


def make_data(time):
    global grids_dong
    '''
    시간대 별 유동인구를 그리드 형태로 반환
    :param time: 0 = morning, 1 = afternoon, ...
    :return: result : 유동인구를 포함한 그리드 numpy 배열
    '''
    print(f'{TIME_DICT[time]} cases to grid...')
    result = np.zeros(shape=(HEIGHT, WIDTH), dtype=int)
    for i in range(15):
        sum = 0
        # i 동, time 시간에서의 유동인구를 가져옴
        case_num = FLOAT_ARR[i][time]
        # i 동의 그리드 안에서
        l = len(grids_dong[i])
        for j in range(int(case_num/10)):
            n = random.randrange(0, l)
            y, x = grids_dong[i][n]
            result[y][x] += 10
            sum += 10
    return result


# 시간대별, 그리드별 유동인구 정보 포함한 배열 생성
grid_time = []
for i in range(4):
    grid_time.append(make_data(i))


# 엑셀 파일 생성(연도별 시트 생성)
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = "Grid"
sheet['B1'] = "Time"
sheet['C1'] = "Float"

# 엑셀 파일에 해당 정보 기록
r = 1
for idx in range(len(grids)):
    y, x = grids[idx]
    for t in range(4):
        sheet.cell(row=r, column=1).value = idx
        sheet.cell(row=r, column=2).value = TIME_DICT[t]
        sheet.cell(row=r, column=3).value = grid_time[t][y][x]
        r += 1

# 엑셀 파일 저장
wb.save('floating_grid' + '.xlsx')
print('Excel file created!')



'''
<동별로 분리된 이미지 생성에 사용>
img = np.zeros(shape=(15, 1, 3), dtype=np.uint8)
for i in range(15):
    img[i] = COLOR_DICT[i]

result = cv2.resize(dsize=(150, 150), src=img, interpolation=cv2.INTER_NEAREST_EXACT)
cv2.imshow('r', result)
cv2.waitKey()

# 그리드 이미지를 배열로 변환
img = cv2.imread("222.png", 1)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)
grids = []

for x in range(WIDTH):
    for y in range(HEIGHT):
        grids.append(img[y][x])

new_list = []
for i in grids:
    alreadyIn = False
    for j in new_list:
        if np.array_equal(i, j):
            alreadyIn = True
    if alreadyIn is False:
        new_list.append(i)
print(new_list)
'''