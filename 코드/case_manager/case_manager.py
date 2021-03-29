import numpy as np
import cv2
import random
import openpyxl


''' ↓ 저장할 엑셀 파일 이름 입력'''
FILE_NAME = "crime_data_x"

'''
O_LATITUDE = 37.5727023
O_LONGTITUDE = 126.961014
LAT_GAP = 0.0008914
LONG_GAP = 0.0011314
'''

O_LATITUDE = 37.572702474
O_LONGTITUDE = 126.960656162
LAT_GAP = 0.000896694
LONG_GAP = 0.001138112

WIDTH = 59
HEIGHT = 32

CRIME_KEY = {0: '절도', 1: '폭행', 2: '살인', 3: '강간', 4: '강도'}
TIME_KEY = {0: '00-03', 1: '03-06', 2: '06-09', 3: '09-12', 4: '12-15', 5: '15-18', 6: '18-21', 7: '21-24' }
YEAR_KEY = {0: '2014', 1: '2015', 2: '2016', 3: '2017', 4: '2018', 5: '2019'}
CRIME_TOTAL = [5231, 4954, 4584, 4184, 4030, 4327]
CRIME_STATS = [
    # 절도    폭행     살인     강간     강도
    [0.4924, 0.9541, 0.9552, 0.9975, 1.0000],  # 2014
    [0.5143, 0.9633, 0.9639, 0.9982, 1.0000],  # 2015
    [0.4690, 0.9544, 0.9551, 0.9983, 1.0000],  # 2016
    [0.4412, 0.9338, 0.9340, 0.9978, 1.0000],  # 2017
    [0.4602, 0.9454, 0.9459, 0.9973, 1.0000],  # 2018
    [0.5089, 0.9531, 0.9535, 0.9986, 1.0000]  # 2019
]
CRIME_TIME = [
    # 00-03  03-06   06-09   09-12   12-15   15-18   18-21   21-24
    [0.0791, 0.1648, 0.2440, 0.3740, 0.5383, 0.7110, 0.8756, 1.0000],  # 절도
    [0.1387, 0.2568, 0.3242, 0.4077, 0.4980, 0.6046, 0.7539, 1.0000],  # 폭행
    [0.0731, 0.1811, 0.2557, 0.3653, 0.4932, 0.6347, 0.8006, 1.0000],  # 살인
    [0.1367, 0.2912, 0.3840, 0.4644, 0.5533, 0.6633, 0.8046, 1.0000],  # 강간
    [0.1574, 0.3538, 0.4198, 0.5082, 0.6102, 0.7181, 0.8396, 1.0000]  # 강도
]
MONTH_DAYS = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
              7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}


# 그리드 이미지를 배열로 변환
img = cv2.imread("grid.png", 0)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)


'''그리드 순회 순서 좌측하단부터 위쪽으로 0번부터 시작!'''
# 배경이 아닌 그리드의 좌표만 배열에 따로 저장 [위도, 경도]
grids = []
result = np.zeros((HEIGHT, WIDTH, 3), dtype=np.uint8)
for i in range(WIDTH):
    for j in range(HEIGHT):
        if img[j][i] != 255:
            grids.append([j, i])


# 엑셀 파일 생성(연도별 시트 생성)
wb = openpyxl.Workbook()
for year in range(6):
    print(f'{YEAR_KEY[year]}...')
    sheet = wb.create_sheet(f'{YEAR_KEY[year]}')
    sheet['A1'] = "Type"
    sheet['B1'] = "Date-Time"
    sheet['C1'] = "Latitude"
    sheet['D1'] = "Longitude"
    sheet['E1'] = "Grid"
    case_num = CRIME_TOTAL[year]

    # 통계자료에 맞춰 범죄사례를 생성
    for i in range(2, case_num + 2):
        # 범죄유형
        n = random.random()
        for c in range(5):
            if n < CRIME_STATS[year][c]:
                sheet.cell(row=i, column=1).value = f'{CRIME_KEY[c]}'
                crime = c
                break

        # 범죄날짜 및 시간
        month = random.randrange(1, 13)
        day = random.randrange(1, MONTH_DAYS[month] + 1)
        date = f'{YEAR_KEY[year]}-{month}-{day}'

        n = random.random()
        for h in range(8):
            if n < CRIME_TIME[crime][h]:
                hour = random.randrange(3*h, 3*h+4)
                break
        time = f'{hour}' + ":" + str(random.randrange(0, 60)) + ":" + str(random.randrange(0, 60))
        sheet.cell(row=i, column=2).value = date + " " + time

        # 범죄지점 : 그리드를 랜덤으로 선택하고, 그리드 내부에서 랜덤생성
        idx = int(random.uniform(0, len(grids)))
        x, y = grids[idx]
        gx, gy = O_LATITUDE - x * LAT_GAP, O_LONGTITUDE + y * LONG_GAP
        sheet.cell(row=i, column=3).value = random.uniform(gx - LAT_GAP, gx)
        sheet.cell(row=i, column=4).value = random.uniform(gy, gy + LONG_GAP)
        sheet.cell(row=i, column=5).value = idx


# 파일 저장
wb.remove(wb['Sheet'])
wb.save(FILE_NAME + '.xlsx')
print('Excel file created!')


